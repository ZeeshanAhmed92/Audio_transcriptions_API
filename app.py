import os
import re
import io
import gc
import sys
import json
import copy
import time
import uuid
import shutil
import queue
import threading
import pandas as pd
import traceback
from dotenv import load_dotenv
from flask import Flask, request, jsonify, send_file, render_template, send_from_directory, make_response, render_template_string
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Alignment, PatternFill, Protection
from utils.transciptions import (generate_html, generate_report, split_and_upload, clean_and_parse_json, process_audio_with_gemini,
                                 export_report_to_excel, generate_improvement_summary) # Added new import

load_dotenv()

# Set base folders and environment variables
os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
app = Flask(__name__)

app.config["JWT_SECRET_KEY"] = os.getenv("JWT_SECRET_KEY")
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = int(os.getenv("JWT_ACCESS_TOKEN_EXPIRES", 3600))
app.config["JWT_TOKEN_LOCATION"] = ["cookies"]
app.config["JWT_COOKIE_NAME"] = "access_token_cookie"
app.config["JWT_COOKIE_SECURE"] = False  # True if using HTTPS
app.config["JWT_COOKIE_CSRF_PROTECT"] = False  # optional for simplicity

UPLOAD_FOLDER = 'uploads'
REPORT_FOLDER = 'reports'
PROJECT_ID = os.getenv("GOOGLE_CLOUD_PROJECT")
LOCATION = "us-central1"
GCS_BUCKET_NAME = "bengali-transcription-bucket1"
TEMP_DIR = "temp_uploads"
GCS_FOLDER_NAME = "Interview_audios"
REPORTS_FOLDER = "reports"
JOBS_STATUS_FILE = "jobs_status.json"
JOBS_FILE = "jobs.json"

# Set up GCP credentials from environment variable
credentials_json = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
if credentials_json and credentials_json.strip().startswith("{"):
    creds_path = "/tmp/gcp-key.json"
    creds_dir = os.path.dirname(creds_path)
    os.makedirs(creds_dir, exist_ok=True)

    with open(creds_path, "w") as f:
        f.write(credentials_json)

    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = creds_path

# Create necessary directories
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORT_FOLDER, exist_ok=True)
os.makedirs(GCS_FOLDER_NAME, exist_ok=True)
os.makedirs(TEMP_DIR, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['REPORT_FOLDER'] = REPORT_FOLDER

# Add shared state for cancellation at the top of your file
jobs_to_cancel = set()
cancellation_lock = threading.Lock()

jwt = JWTManager(app)

# --- Persistent state for jobs_status ---
def read_job_statuses():
    """Reads job statuses from a JSON file."""
    try:
        if os.path.exists(JOBS_STATUS_FILE):
            with open(JOBS_STATUS_FILE, "r") as f:
                return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        pass
    return {}


def write_job_statuses(statuses):
    """Writes job statuses to a JSON file."""
    with open(JOBS_STATUS_FILE, "w") as f:
        json.dump(statuses, f, indent=4)


# Initialize job statuses from the file at startup
jobs_status = read_job_statuses()
report_queue = queue.Queue()


def read_jobs():
    if os.path.exists(JOBS_FILE):
        with open(JOBS_FILE, "r") as f:
            return json.load(f)
    return []


def write_jobs(jobs):
    with open(JOBS_FILE, "w") as f:
        json.dump(jobs, f, indent=4)


def delete_job_folders(job_id):
    report_path = os.path.join(REPORT_FOLDER, f"job_{job_id}")
    upload_path = os.path.join(UPLOAD_FOLDER, f"job_{job_id}")

    for path in [report_path, upload_path]:
        if os.path.exists(path):
            shutil.rmtree(path)


@app.route("/all_jobs_status")
@jwt_required()
def all_jobs_status():
    return jsonify(jobs_status)


@app.route("/get_questions", methods=["GET"])
@jwt_required()
def get_questions():
    questions_dir = "./utils/questionaires"
    if not os.path.exists(questions_dir):
        return jsonify([])

    files = [
        f for f in os.listdir(questions_dir)
        if os.path.isfile(os.path.join(questions_dir, f))
    ]
    return jsonify(files)

@app.route("/delete_job/<int:job_id>", methods=["DELETE"])
@jwt_required()
def delete_job(job_id):
    """
    Deletes all files and folders associated with a specific job_id,
    and also signals the background worker to cancel any in-progress
    tasks related to that job.
    """
    job_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")
    job_report_folder = os.path.join(app.config['REPORT_FOLDER'], f"job_{job_id}")

    # Step 1: Find all related jobs in the queue and flag them for cancellation.
    jobs_to_remove = []
    with cancellation_lock:
        for jid, status in list(jobs_status.items()):
            # Check if the job's folder path matches the job_id
            if str(status.get("job_folder", "")).endswith(f"job_{job_id}"):
                jobs_to_cancel.add(jid)  # Signal the worker to stop this specific job
                jobs_to_remove.append(jid)
    
    # Step 2: Remove the job entries from the in-memory status dictionary.
    for jid in jobs_to_remove:
        jobs_status.pop(jid, None)

    # Step 3: Delete the physical folders from the file system.
    removed_any = False
    try:
        if os.path.exists(job_upload_folder):
            shutil.rmtree(job_upload_folder)
            removed_any = True
        if os.path.exists(job_report_folder):
            shutil.rmtree(job_report_folder)
            removed_any = True
    except OSError as e:
        print(f"Error deleting job folders: {e}")
        # The job should still be removed from the list even if files can't be deleted.
        pass

    # Step 4: Remove the job from the main jobs.json file.
    jobs = read_jobs()
    original_job_count = len(jobs)
    jobs = [job for job in jobs if job["id"] != job_id]
    
    if len(jobs) < original_job_count:
        write_jobs(jobs)
        write_job_statuses(jobs_status) # Save the updated in-memory job statuses
        return jsonify({"msg": f"Job {job_id} and all related data have been successfully deleted."}), 200

    # If the job was not found in jobs.json, return a not found message.
    return jsonify({"msg": "Job not found"}), 404


def report_worker():
    while True:
        job_id, filename, job_number, questionaire = report_queue.get()
        try:
            # Check for cancellation at the start
            with cancellation_lock:
                if job_id in jobs_to_cancel:
                    print(f"[Worker] Job {job_id} was cancelled. Skipping.")
                    jobs_to_cancel.remove(job_id)
                    jobs_status[job_id]["status"] = "cancelled"
                    jobs_status[job_id]["error"] = "Job was deleted by user."
                    write_job_statuses(jobs_status)
                    report_queue.task_done()
                    continue

            if job_id not in jobs_status:
                print(f"[Worker] Warning: job_id {job_id} not found in jobs_status")
                report_queue.task_done()
                continue

            jobs_status[job_id]["status"] = "Processing"
            write_job_statuses(jobs_status)
            print(f"[Worker] Start job_id={job_id}, file={filename}, job_number={job_number}")

            # --- Folders and Paths ---
            audio_path = os.path.join(UPLOAD_FOLDER, f"job_{job_number}", filename)
            if not os.path.exists(audio_path):
                raise FileNotFoundError(f"Audio file not found: {audio_path}")

            report_job_folder = os.path.join(REPORT_FOLDER, f"job_{job_number}")
            os.makedirs(report_job_folder, exist_ok=True)

            base_name = os.path.splitext(filename)[0]
            meta_path = os.path.join(report_job_folder, f"{base_name}_meta.json")
            transcription_path = os.path.join(report_job_folder, f"{base_name}_transcription.json")
            report_json_path = os.path.join(report_job_folder, f"{base_name}_evaluation_report.json")
            report_excel_path = os.path.join(report_job_folder, f"{base_name}_evaluation_report.xlsx")
            report_html_path = os.path.join(report_job_folder, f"{base_name}_evaluation_report.html")
            report_summary_html_path = os.path.join(report_job_folder, f"{base_name}_summary.html") # New summary path

            meta = {
                "gcs_uri": None,
                "transcription_done": False,
                "report_done": False,
                "questionaire": questionaire
            }
            if os.path.exists(meta_path):
                with open(meta_path, "r") as f:
                    meta = json.load(f)

            # Reset meta if questionnaire changes
            if meta.get("questionaire") != questionaire:
                meta = {
                    "gcs_uri": None,
                    "transcription_done": False,
                    "report_done": False,
                    "questionaire": questionaire
                }

            # --- Step 1: Upload to GCS ---
            if not meta.get("gcs_uri"):
                jobs_status[job_id]["status"] = "Uploading to GCS"
                write_job_statuses(jobs_status)
                print(f"[Worker] Uploading {filename} to GCS...")
                gcs_uris = split_and_upload(GCS_BUCKET_NAME, audio_path, f"{GCS_FOLDER_NAME}/{filename}")
                meta["gcs_uri"] = gcs_uris
                with open(meta_path, "w") as f:
                    json.dump(meta, f, indent=4)
            else:
                print("[Worker] GCS upload already done.")

            # Check for cancellation before transcription
            with cancellation_lock:
                if job_id in jobs_to_cancel:
                    print(f"[Worker] Job {job_id} cancelled during GCS upload. Skipping.")
                    jobs_to_cancel.remove(job_id)
                    jobs_status[job_id]["status"] = "cancelled"
                    jobs_status[job_id]["error"] = "Job was deleted by user."
                    write_job_statuses(jobs_status)
                    report_queue.task_done()
                    continue

            # --- Step 2: Transcription ---
            if not meta.get("transcription_done") or not os.path.exists(transcription_path):
                jobs_status[job_id]["status"] = "Transcribing Audio"
                write_job_statuses(jobs_status)
                print(f"[Worker] Transcribing {filename}...")

                all_transcriptions = []
                for uri in meta["gcs_uri"]:
                    raw_transcription = process_audio_with_gemini(PROJECT_ID, LOCATION, uri)
                    parsed = clean_and_parse_json(raw_transcription)
                    if parsed:
                        all_transcriptions.extend(parsed)

                if all_transcriptions:
                    with open(transcription_path, "w", encoding="utf-8") as f:
                        json.dump(all_transcriptions, f, ensure_ascii=False, indent=4)
                    meta["transcription_done"] = True
                    with open(meta_path, "w") as f:
                        json.dump(meta, f, indent=4)
                else:
                    jobs_status[job_id]["status"] = "error"
                    jobs_status[job_id]["error"] = "Transcription failed"
                    write_job_statuses(jobs_status)
                    continue
            else:
                print("[Worker] Transcription already done.")

            # Check for cancellation before report generation
            with cancellation_lock:
                if job_id in jobs_to_cancel:
                    print(f"[Worker] Job {job_id} cancelled during transcription. Skipping.")
                    jobs_to_cancel.remove(job_id)
                    jobs_status[job_id]["status"] = "cancelled"
                    jobs_status[job_id]["error"] = "Job was deleted by user."
                    write_job_statuses(jobs_status)
                    report_queue.task_done()
                    continue

            # --- Step 3: Report Generation ---
            if not meta.get("report_done") or not os.path.exists(report_excel_path):
                jobs_status[job_id]["status"] = "Generating Report"
                write_job_statuses(jobs_status)
                print(f"[Worker] Generating report for {filename}...")

                with open(transcription_path, "r", encoding="utf-8") as f:
                    transcription_json = json.load(f)

                report_text = generate_report(json.dumps(transcription_json), questionaire)
                report_json = clean_and_parse_json(report_text)

                if report_json:
                    with open(report_json_path, "w", encoding="utf-8") as f:
                        json.dump(report_json, f, ensure_ascii=False, indent=4)
                    
                    # Generate all report formats
                    generate_html(report_json, report_html_path)
                    export_report_to_excel(report_json, report_excel_path)
                    # ✅ Generate the new summary HTML
                    summary_html_content = generate_improvement_summary(report_json)
                    with open(report_summary_html_path, "w", encoding="utf-8") as f:
                        f.write(summary_html_content)

                    meta["report_done"] = True
                    with open(meta_path, "w") as f:
                        json.dump(meta, f, indent=4)
                else:
                    jobs_status[job_id]["status"] = "error"
                    jobs_status[job_id]["error"] = "Report generation failed"
                    write_job_statuses(jobs_status)
                    continue
            else:
                print("[Worker] Report already generated.")

            # --- Mark job as done ---
            jobs_status[job_id]["status"] = "done"
            jobs_status[job_id]["report_excel"] = os.path.basename(report_excel_path)
            jobs_status[job_id]["report_html"] = os.path.basename(report_html_path)
            jobs_status[job_id]["report_summary_html"] = os.path.basename(report_summary_html_path) # New path
            write_job_statuses(jobs_status)

        except Exception as e:
            print(f"[Worker] Error: {e}")
            traceback.print_exc()
            if job_id in jobs_status:
                jobs_status[job_id]["status"] = "error"
                jobs_status[job_id]["error"] = str(e)
                write_job_statuses(jobs_status)
        finally:
            report_queue.task_done()

threading.Thread(target=report_worker, daemon=True).start()

@app.route("/generate_report", methods=["POST"])
@jwt_required()
def generate_report_route():
    data = request.json
    filename = data.get("filename")
    questionaire = data.get("questionnaire")
    job_number = str(data.get("job_id"))  # keep consistent naming

    if not filename or not job_number or not questionaire:
        return jsonify({"error": "Filename, questionnaire, or job_id missing"}), 400

    job_upload_folder = os.path.join(UPLOAD_FOLDER, f"job_{job_number}")
    if not os.path.exists(os.path.join(job_upload_folder, filename)):
        return jsonify({"error": "Audio file not found"}), 404

    report_folder = os.path.join(REPORT_FOLDER, f"job_{job_number}")
    os.makedirs(report_folder, exist_ok=True)

    # Generate unique job_id for queue tracking
    job_id = str(uuid.uuid4())
    base_name = os.path.splitext(filename)[0]

    meta_path = os.path.join(report_folder, f"{base_name}_meta.json")
    report_excel_path = os.path.join(report_folder, f"{base_name}_evaluation_report.xlsx")
    report_html_path = os.path.join(report_folder, f"{base_name}_evaluation_report.html")
    report_summary_html_path = os.path.join(report_folder, f"{base_name}_summary.html") # New path

    # If meta exists and questionnaire matches -> reuse
    if os.path.exists(meta_path) and os.path.exists(report_excel_path) and os.path.exists(report_html_path):
        with open(meta_path, "r") as f:
            meta = json.load(f)
        if meta.get("questionaire") == questionaire:
            jobs_status[job_id] = {
                "status": "done",
                "job_id": job_id,
                "file": filename,
                "job_folder": report_folder,
                "report_excel": os.path.basename(report_excel_path),
                "report_html": os.path.basename(report_html_path),
                "report_summary_html": os.path.basename(report_summary_html_path) # New path
            }
            write_job_statuses(jobs_status)
            return jsonify({
                "job_id": job_id,
                "status": "done",
                "report_excel": os.path.basename(report_excel_path),
                "report_html": os.path.basename(report_html_path),
                "report_summary_html": os.path.basename(report_summary_html_path) # New path
            })

    # Otherwise -> enqueue fresh processing
    jobs_status[job_id] = {"status": "pending", "file": filename, "job_folder": report_folder}
    write_job_statuses(jobs_status)
    report_queue.put((job_id, filename, job_number, questionaire))
    return jsonify({"job_id": job_id, "status": "queued"})


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/merged_report', methods=['POST'])
@jwt_required()
def merge_reports():
    data = request.get_json()
    job_id = data.get("job_id")
    filenames = data.get("filenames", [])

    if not job_id or not filenames:
        return jsonify({"msg": "job_id and filenames are required"}), 400

    job_report_folder = os.path.join(app.config['REPORT_FOLDER'], f"job_{job_id}")
    if not os.path.exists(job_report_folder):
        return jsonify({"msg": f"Job {job_id} report folder not found"}), 404

    merged_wb = Workbook()
    merged_ws_report = merged_wb.active
    merged_ws_report.title = "Merged Report"

    merged_ws_summary = merged_wb.create_sheet(title="Merged Summary")

    current_row_report = 1
    current_row_summary = 1

    for filename in filenames:
        report_path = os.path.join(job_report_folder, filename)

        if not os.path.exists(report_path):
            return jsonify({"msg": f"Report not found: {filename}"}), 404

        try:
            wb = load_workbook(report_path)

            # Merge "Report" sheet
            if "Report" in wb.sheetnames:
                ws = wb["Report"]
                merged_ws_report.cell(row=current_row_report, column=1, value=f"Report: {filename}")
                merged_ws_report.cell(row=current_row_report, column=1).font = Font(bold=True)
                current_row_report += 1

                for row in ws.iter_rows():
                    for col_index, cell in enumerate(row, start=1):
                        new_cell = merged_ws_report.cell(row=current_row_report, column=col_index, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.alignment = copy.copy(cell.alignment)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy.copy(cell.protection)
                    current_row_report += 1

                current_row_report += 1

            # Merge "Summary" sheet
            if "Summary" in wb.sheetnames:
                ws_summary = wb["Summary"]
                merged_ws_summary.cell(row=current_row_summary, column=1, value=f"Summary: {filename}")
                merged_ws_summary.cell(row=current_row_summary, column=1).font = Font(bold=True)
                current_row_summary += 1

                for row in ws_summary.iter_rows():
                    for col_index, cell in enumerate(row, start=1):
                        new_cell = merged_ws_summary.cell(row=current_row_summary, column=col_index, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy.copy(cell.font)
                            new_cell.fill = copy.copy(cell.fill)
                            new_cell.border = copy.copy(cell.border)
                            new_cell.alignment = copy.copy(cell.alignment)
                            new_cell.number_format = cell.number_format
                            new_cell.protection = copy.copy(cell.protection)
                    current_row_summary += 1

                current_row_summary += 1

        except Exception as e:
            return jsonify({"msg": f"Error reading {filename}: {str(e)}"}), 500

    merged_path = os.path.join(job_report_folder, f"merged_{job_id}.xlsx")
    merged_wb.save(merged_path)

    return send_file(merged_path, as_attachment=True)


@app.route('/merged_html_report', methods=['POST'])
@jwt_required()
def merge_html_reports():
    data = request.get_json()
    job_id = data.get("job_id")
    filenames = data.get("filenames", [])

    if not job_id or not filenames:
        return jsonify({"msg": "job_id and filenames are required"}), 400

    job_report_folder = os.path.join(app.config['REPORT_FOLDER'], f"job_{job_id}")
    if not os.path.exists(job_report_folder):
        return jsonify({"msg": f"Job {job_id} report folder not found"}), 404

    merged_content = ""
    for filename in filenames:
        report_path = os.path.join(job_report_folder, filename)

        if not os.path.exists(report_path):
            return jsonify({"msg": f"Report not found: {filename}"}), 404

        try:
            with open(report_path, 'r', encoding='utf-8') as f:
                content = f.read()

            merged_content += f"""
            <div class="report-section">
                <h2>{filename}</h2>
                {content}
            </div>
            """
        except Exception as e:
            return jsonify({"msg": f"Error reading {filename}: {str(e)}"}), 500

    html_template = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Merged Report</title>
        <style>
            body { font-family: sans-serif; line-height: 1.6; margin: 20px; }
            .report-section {
                border: 1px solid #ccc;
                padding: 20px;
                margin-bottom: 20px;
                border-radius: 8px;
                background: #fafafa;
            }
            h1, h2 { color: #333; }
        </style>
    </head>
    <body>
        <h1>Merged Evaluation Report</h1>
        {{ merged_content | safe }}
    </body>
    </html>
    """
    final_html = render_template_string(html_template, merged_content=merged_content)

    merged_path = os.path.join(job_report_folder, f"merged_{job_id}.html")
    with open(merged_path, "w", encoding="utf-8") as f:
        f.write(final_html)

    return send_file(merged_path, mimetype="text/html", as_attachment=True)


@app.route("/merged_summary_html", methods=["POST"])
@jwt_required()
def merged_summary_html():
    data = request.get_json()
    job_id = data.get("job_id")
    filenames = data.get("filenames", [])

    if not job_id or not filenames:
        return jsonify({"error": "Missing job_id or filenames"}), 400

    job_folder = os.path.join(app.config["REPORT_FOLDER"], f"job_{job_id}")
    merged_path = os.path.join(job_folder, f"merged_summary_{job_id}.html")

    merged_html = "<html><body>"
    for filename in filenames:
        file_path = os.path.join(job_folder, filename)  # summary.html stored in same folder
        if os.path.exists(file_path):
            with open(file_path, "r", encoding="utf-8") as f:
                merged_html += f"<h2>{filename}</h2>\n"
                merged_html += f.read()
                merged_html += "<hr/>"

    merged_html += "</body></html>"

    with open(merged_path, "w", encoding="utf-8") as f:
        f.write(merged_html)

    return send_file(
        merged_path,
        mimetype="text/html",
        as_attachment=True,
        download_name=os.path.basename(merged_path),
    )


@app.route("/report/<int:job_id>")
@jwt_required()
def report_page(job_id):
    current_user = get_jwt_identity()
    # Ensure the job exists
    with open(JOBS_FILE) as f:
        jobs = json.load(f)
    job = next((j for j in jobs if j["id"] == job_id), None)
    if not job:
        return "Job not found", 404

    # Create job-specific upload folder if not exists
    job_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")
    os.makedirs(job_folder, exist_ok=True)

    return render_template("report.html", job_id=job_id, user=current_user)


# Login endpoint
@app.route("/login", methods=["POST"])
def login():
    data = request.json
    username = data.get("username")
    password = data.get("password")

    if username == os.getenv('APP_USERNAME') and password == os.getenv('APP_PASSWORD'):
        access_token = create_access_token(identity=username)
        resp = make_response(jsonify({"msg": "Login successful"}))
        resp.set_cookie(
            "access_token_cookie",
            access_token,
            httponly=True,
            samesite="Lax"
        )
        return resp

    return jsonify({"msg": "Bad username or password"}), 401


@app.route("/create_job", methods=["POST"])
@jwt_required()
def create_job():
    data = request.json
    job_name = data.get("job_name")
    if not job_name:
        return jsonify({"msg": "Job name required"}), 400

    # Load existing jobs
    if os.path.exists(JOBS_FILE):
        with open(JOBS_FILE) as f:
            jobs = json.load(f)
    else:
        jobs = []

    job_id = len(jobs) + 1
    job = {"id": job_id, "name": job_name}
    jobs.append(job)

    # Save jobs
    with open(JOBS_FILE, "w") as f:
        json.dump(jobs, f, indent=4)

    # ✅ Create job-specific folders right away
    job_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")
    job_report_folder = os.path.join(app.config['REPORT_FOLDER'], f"job_{job_id}")
    os.makedirs(job_upload_folder, exist_ok=True)
    os.makedirs(job_report_folder, exist_ok=True)

    return jsonify({"msg": "Job created", "job": job})


# Protected route
@app.route("/protected")
@jwt_required()
def protected():
    current_user = get_jwt_identity()
    return jsonify(logged_in_as=current_user)

# Logout
@app.route("/logout", methods=["POST"])
def logout():
    resp = make_response(jsonify({"msg": "Logged out"}))
    resp.delete_cookie("access_token_cookie")
    return resp


@app.route('/upload/<int:job_id>', methods=['POST'])
@jwt_required()
def upload_audio(job_id):
    if 'audio_file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['audio_file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    job_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")
    os.makedirs(job_folder, exist_ok=True)

    filepath = os.path.join(job_folder, file.filename)
    file.save(filepath)
    return jsonify({"message": "File uploaded successfully"})


@app.route('/list_files/<int:job_id>', methods=['GET'])
@jwt_required()
def list_files(job_id):
    job_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")
    if not os.path.exists(job_folder):
        return jsonify([])

    files = [
        f for f in os.listdir(job_folder)
        if os.path.isfile(os.path.join(job_folder, f))
    ]

    return jsonify(files)

@app.route("/report_status/<job_id>")
@jwt_required()
def report_status(job_id):
    job = jobs_status.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    report_folder = job.get("job_folder")
    filename = job.get("file")
    base_name = os.path.splitext(filename)[0]

    # Expected report file names
    report_excel_path = os.path.join(report_folder, f"{base_name}_evaluation_report.xlsx")
    report_html_path = os.path.join(report_folder, f"{base_name}_evaluation_report.html")
    report_summary_html_path = os.path.join(report_folder, f"{base_name}_summary.html") # New path

    # If status is done but paths not stored yet -> update jobs_status
    if job["status"] == "done":
        if os.path.exists(report_excel_path):
            job["report_excel"] = os.path.basename(report_excel_path)
        if os.path.exists(report_html_path):
            job["report_html"] = os.path.basename(report_html_path)
        if os.path.exists(report_summary_html_path): # New path
            job["report_summary_html"] = os.path.basename(report_summary_html_path)
        jobs_status[job_id] = job
        write_job_statuses(jobs_status)

    return jsonify(job)

@app.route("/download_report/<job_id>/<filename>", methods=["GET"])
def download_report(job_id, filename):
    """Download an Excel report for a specific job."""
    job_folder = os.path.join(app.config["REPORT_FOLDER"], f"job_{job_id}")
    file_path = os.path.join(job_folder, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "Report not found"}), 404
    return send_from_directory(job_folder, filename, as_attachment=True)


@app.route("/download_html_report/<job_id>/<filename>", methods=["GET"])
def download_html_report(job_id, filename):
    """Download an HTML report for a specific job."""
    job_folder = os.path.join(app.config["REPORT_FOLDER"], f"job_{job_id}")
    file_path = os.path.join(job_folder, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "Report not found"}), 404
    return send_from_directory(job_folder, filename, as_attachment=True)

@app.route("/download_summary_report/<job_id>/<filename>", methods=["GET"])
def download_summary_report(job_id, filename):
    """Download the summary HTML report for a specific job."""
    job_folder = os.path.join(app.config["REPORT_FOLDER"], f"job_{job_id}")
    file_path = os.path.join(job_folder, filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "Summary report not found"}), 404
    return send_from_directory(job_folder, filename, as_attachment=True)


@app.route("/delete_files", methods=["DELETE"])
@jwt_required()
def delete_files():
    data = request.get_json()
    job_id_folder = data.get("job_id")
    files = data.get("files", [])
    if not job_id_folder or not files:
        return jsonify({"msg": "job_id and files are required"}), 400

    job_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id_folder}")
    job_report_folder = os.path.join(app.config['REPORT_FOLDER'], f"job_{job_id_folder}")

    deleted, not_found = [], []
    jobs_to_remove = []

    for filename in files:
        base_name = os.path.splitext(filename)[0]

        # 1. Delete audio file
        file_path = os.path.join(job_upload_folder, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            deleted.append(filename)
        else:
            not_found.append(filename)

        # 2. Delete all related report files
        if os.path.exists(job_report_folder):
            for f in os.listdir(job_report_folder):
                if base_name in f:
                    os.remove(os.path.join(job_report_folder, f))

        # 3. Find job IDs and add them to the cancellation set
        for jid, status in list(jobs_status.items()):
            if str(status.get("job_folder", "")).endswith(f"job_{job_id_folder}") and status.get("file") == filename:
                with cancellation_lock:
                    jobs_to_cancel.add(jid)
                jobs_to_remove.append(jid)

    for jid in jobs_to_remove:
        jobs_status.pop(jid, None)

    write_job_statuses(jobs_status)

    # Check if the job folder is empty and delete the job if it is
    if os.path.exists(job_upload_folder) and not os.listdir(job_upload_folder):
        delete_job_folders(job_id_folder)
        
        # Remove the job from jobs.json as well
        jobs = read_jobs()
        jobs = [job for job in jobs if job["id"] != job_id_folder]
        write_jobs(jobs)

    return jsonify({
        "deleted": deleted,
        "not_found": not_found,
        "msg": f"Files {deleted} deleted successfully. In-progress jobs will be cancelled."
    }), 200


@app.route("/jobs")
@jwt_required()
def list_jobs():
    current_user = get_jwt_identity()
    with open(JOBS_FILE) as f:
        jobs = json.load(f)

    jobs_with_counts = []
    for job in jobs:
        job_id = job["id"]
        job_folder = os.path.join(app.config['UPLOAD_FOLDER'], f"job_{job_id}")

        # Count total audio files
        total_audios = len([
            f for f in os.listdir(job_folder)
            if f.lower().endswith(".wav")
        ]) if os.path.exists(job_folder) else 0

        # Count total runs in queue or processing for this job_id
        active_runs = sum(
            1 for item in list(report_queue.queue)
            if str(item[2]) == str(job_id)
        ) + sum(
            1 for jid, status in jobs_status.items()
            if str(status.get("job_folder", "")).endswith(f"job_{job_id}")
            and status.get("status") in ["Pending", "Processing", "Uploading to GCS", "Transcribing Audio", "Generating Report"]
        )

        jobs_with_counts.append({
            "id": job_id,
            "name": job["name"],
            "total_audios": total_audios,
            "active_runs": active_runs,
            "current_user": current_user
        })

    return render_template("jobs.html", jobs=jobs_with_counts)


if __name__ == '__main__':
    os.makedirs(GCS_FOLDER_NAME, exist_ok=True)
    os.makedirs(TEMP_DIR, exist_ok=True)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(REPORT_FOLDER, exist_ok=True)
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)